"""Generate Lab 9 growth-projection.xlsx for Bandersnatch."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

LTV = 5.10
MONTHS = ["M1", "M2", "M3", "M4", "M5", "M6"]
HEADERS = [
    "Month",
    "Channel",
    "Spend (USD)",
    "Visitors",
    "Signups",
    "Activated",
    "Retained D30",
    "CAC (USD)",
    "LTV (USD)",
    "LTV:CAC",
    "Cumulative Users",
]

# Per-channel monthly inputs (Expected case)
QR_SPEND = [5.60] * 6
QR_VISITORS = [250, 288, 317, 333, 350, 368]
QR_SIGNUP_RATE = 0.22

GROUPS_SPEND = [0.0] * 6
GROUPS_VISITORS = [800, 400, 200, 100, 80, 60]
GROUPS_SIGNUP_RATE = 0.05

REFERRAL_SPEND = [0.0] * 6
REFERRAL_VISITORS = [0, 25, 50, 75, 100, 130]
REFERRAL_SIGNUP_RATE = 0.25

D30_RETENTION = [0.50, 0.51, 0.52, 0.53, 0.54, 0.55]

ACTIVATION = {"Expected Case": 0.31, "Best Case": 0.465, "Worst Case": 0.155}

SCENARIO_NOTE = (
    "Scenario variable: Signup-to-activation rate. "
    "Expected 31%, Best 46.5% (+50%), Worst 15.5% (-50%). "
    "All other inputs identical across scenarios."
)


def build_rows(activation_rate: float) -> list[dict]:
    rows = []
    cumulative = 0.0
    channels = [
        ("QR Flyers", QR_SPEND, QR_VISITORS, QR_SIGNUP_RATE),
        ("Student Groups", GROUPS_SPEND, GROUPS_VISITORS, GROUPS_SIGNUP_RATE),
        ("Referral", REFERRAL_SPEND, REFERRAL_VISITORS, REFERRAL_SIGNUP_RATE),
    ]
    for mi, month in enumerate(MONTHS):
        d30 = D30_RETENTION[mi]
        month_retained = 0.0
        for ch_name, spends, visitors, signup_rate in channels:
            v = visitors[mi]
            s = round(v * signup_rate, 1)
            a = round(s * activation_rate, 1)
            r = round(a * d30, 1)
            spend = spends[mi]
            cac = round(spend / r, 2) if r > 0 else None
            ltv_cac = round(LTV / cac, 1) if cac and cac > 0 else None
            cumulative += r
            month_retained += r
            rows.append(
                {
                    "month": month,
                    "channel": ch_name,
                    "spend": spend,
                    "visitors": v,
                    "signups": s,
                    "activated": a,
                    "retained": r,
                    "cac": cac,
                    "ltv": LTV,
                    "ltv_cac": ltv_cac,
                    "cumulative": round(cumulative, 1),
                }
            )
        # Blended row
        total_spend = sum(c[1][mi] for c in channels)
        blended_cac = (
            round(total_spend / month_retained, 2) if month_retained > 0 else None
        )
        blended_ltv_cac = (
            round(LTV / blended_cac, 1) if blended_cac and blended_cac > 0 else None
        )
        rows.append(
            {
                "month": month,
                "channel": "Blended",
                "spend": total_spend,
                "visitors": sum(c[2][mi] for c in channels),
                "signups": round(
                    sum(
                        visitors[mi] * signup_rate
                        for _, _, visitors, signup_rate in channels
                    ),
                    1,
                ),
                "activated": round(
                    sum(
                        visitors[mi] * signup_rate * activation_rate
                        for _, _, visitors, signup_rate in channels
                    ),
                    1,
                ),
                "retained": round(month_retained, 1),
                "cac": blended_cac,
                "ltv": LTV,
                "ltv_cac": blended_ltv_cac,
                "cumulative": round(cumulative, 1),
                "bold": True,
            }
        )
    return rows


def write_scenario_sheet(ws, scenario_name: str, activation_rate: float):
    ws.append([f"{scenario_name} — Bandersnatch Bus #3 Tracker"])
    ws.append([SCENARIO_NOTE])
    ws.append([])
    ws.append(HEADERS)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    rows = build_rows(activation_rate)
    for r in rows:
        ws.append(
            [
                r["month"],
                r["channel"],
                r["spend"],
                r["visitors"],
                r["signups"],
                r["activated"],
                r["retained"],
                r["cac"] if r["cac"] is not None else "N/A",
                r["ltv"],
                r["ltv_cac"] if r["ltv_cac"] is not None else "N/A",
                r["cumulative"],
            ]
        )
        if r.get("bold"):
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def write_assumptions_sheet(ws):
    ws.append(["Assumptions — Bandersnatch Growth Projection"])
    ws.append([])
    ws.append(["Input", "Value used", "Source", "Confidence"])
    for cell in ws[3]:
        cell.font = Font(bold=True)
    data = [
        ("Monetisable value (LTV base)", "$1.00/month", "Interviews P03, P11", "Medium"),
        ("Gross margin", "85%", "Future infra estimate", "Medium"),
        ("Customer lifetime", "6 months", "Academic semester proxy", "Medium"),
        ("LTV", "$5.10", "1.00 x 0.85 x 6", "Calculated"),
        ("QR visitor→signup", "22%", "Landing page proxy [assumption]", "Low"),
        ("Groups visitor→signup", "5%", "Team estimate", "Low"),
        ("Referral invite→signup", "25%", "Landing page proxy", "Low"),
        ("Signup→activation (Expected)", "31%", "Benchmark 25–40%; GA4 TBD", "Low"),
        ("Signup→activation (Best)", "46.5%", "Expected +50%", "Scenario"),
        ("Signup→activation (Worst)", "15.5%", "Expected -50%", "Scenario"),
        ("D30 retention M1→M6", "50%→55%", "Commute app benchmark; improves with product", "Low"),
        ("QR spend/month", "$5.60", "Team budget", "High"),
        ("QR visitors M1", "250", "5 stops x foot-traffic estimate", "Low"),
        ("Organic visitor MoM growth", "15%→5%", "Group chat / stop saturation", "Assumption"),
        ("Groups visitors decay", "800→60", "One-shot posts per group", "Medium"),
        ("Referral visitors", "0→130", "K=0.20; base growth", "Low"),
        ("K-factor", "0.20", "0.8 invites x 25% conversion", "Assumption"),
    ]
    for row in data:
        ws.append(list(row))
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 28


def main():
    out = r"c:\Users\Laptop\Documents\GitHub\product-capstone-2026\04-gtm\growth-projection.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    for name, rate in ACTIVATION.items():
        ws = wb.create_sheet(name)
        write_scenario_sheet(ws, name, rate)
    ws_a = wb.create_sheet("Assumptions")
    write_assumptions_sheet(ws_a)
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
